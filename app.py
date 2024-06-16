import streamlit as st
import pandas as pd
import re
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def process_text(text):
    cleaned_text = re.sub(r'#### \d+\.', '', text)
    matches = re.findall(r'Judul:\s*"([^"]+)"\s*-?\s*Tujuan:\s*([^"]+)', cleaned_text)

    if not matches:
        st.error("Tidak dapat menemukan pasangan Judul dan Tujuan yang sesuai dalam teks yang diberikan.")
        return None

    data = {
        'Judul': [match[0].strip() for match in matches],
        'Tujuan': [match[1].replace('Judul:', '').strip() for match in matches]
    }
    df = pd.DataFrame(data)
    return df

def style_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = ["Judul", "Tujuan"]
    ws.append(headers)

    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    # Ubah warna header jadi hitam
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    for cell in ws["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal="left", vertical="top")

    return wb

def dataframe_to_rows(df, index=True, header=True):
    rows = df.itertuples(index=index, name=None)
    if header:
        yield df.columns.tolist()
    for row in rows:
        yield list(row)

def main():
    st.title('Proses File .txt ke Excel')
    uploaded_file = st.file_uploader("Unggah file .txt", type='txt')

    if uploaded_file is not None:
        st.text("File yang diunggah:")
        st.write(uploaded_file.name)

        if st.button('Proses'):
            text = uploaded_file.read().decode('utf-8')
            df = process_text(text)
            if df is not None:
                st.success("Data berhasil diproses:")
                st.markdown("**Hasil Konversi:**")
                st.dataframe(df)

                wb = style_excel(df)
                output = BytesIO()
                wb.save(output)
                processed_data = output.getvalue()

                st.download_button(
                    label="Download Hasil",
                    data=processed_data,
                    file_name='hasil.xlsx',
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

if __name__ == '__main__':
    main()
